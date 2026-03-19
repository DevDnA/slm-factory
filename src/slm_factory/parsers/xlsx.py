"""XLSX (Excel) 문서 파서입니다."""

from __future__ import annotations

from pathlib import Path
from typing import ClassVar

from ..models import ParsedDocument
from ..utils import get_logger
from .base import BaseParser, extract_date_from_filename, rows_to_markdown

logger = get_logger("parsers.xlsx")


class XLSXParser(BaseParser):
    """XLSX 스프레드시트를 파싱합니다."""

    extensions: ClassVar[list[str]] = [".xlsx"]

    def can_parse_content(self, path: Path) -> bool:
        import zipfile

        try:
            if not zipfile.is_zipfile(str(path)):
                return False
            with zipfile.ZipFile(path, "r") as zf:
                return "xl/workbook.xml" in zf.namelist()
        except Exception:
            return False

    def parse(self, path: Path) -> ParsedDocument:
        """XLSX 파일을 파싱하여 ParsedDocument를 반환합니다.

        각 시트를 마크다운 표로 변환합니다.

        매개변수
        ----------
        path:
            XLSX 파일의 경로입니다.

        반환값
        -------
        ParsedDocument
            텍스트 내용, 표 (마크다운), 메타데이터로 채워집니다.

        예외
        ------
        ImportError
            openpyxl이 설치되지 않았으면 발생합니다.
        FileNotFoundError
            *path*가 존재하지 않으면 발생합니다.
        RuntimeError
            XLSX 파일을 파싱할 수 없으면 발생합니다.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl이 설치되지 않았습니다. uv sync --extra xlsx 로 설치하세요."
            ) from exc

        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"XLSX file not found: {path}")

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise RuntimeError(
                f"XLSX 파일을 파싱할 수 없습니다: {path}\n"
                f"원인: 파일이 손상되었거나 openpyxl이 지원하지 않는 형식일 수 있습니다.\n"
                f"해결: Excel에서 다시 저장하거나 CSV로 변환하세요."
            ) from exc

        # ------------------------------------------------------------------
        # 시트별 텍스트 및 표 추출
        # ------------------------------------------------------------------
        sheet_texts: list[str] = []
        tables_md: list[str] = []

        try:
            for sheet in wb.worksheets:
                rows: list[list[str]] = []
                for row in sheet.iter_rows():
                    cells = [
                        str(cell.value).strip() if cell.value is not None else ""
                        for cell in row
                    ]
                    # 완전히 빈 행 건너뛰기
                    if any(c for c in cells):
                        rows.append(cells)

                if not rows:
                    continue

                md_table = rows_to_markdown(rows)
                if md_table:
                    header = f"## {sheet.title}" if len(wb.worksheets) > 1 else ""
                    if header:
                        sheet_texts.append(f"{header}\n\n{md_table}")
                    else:
                        sheet_texts.append(md_table)
                    tables_md.append(md_table)
        finally:
            wb.close()

        content = "\n\n".join(sheet_texts)

        # ------------------------------------------------------------------
        # 메타데이터
        # ------------------------------------------------------------------
        title = path.stem
        metadata: dict[str, object] = {
            "sheet_count": len(wb.sheetnames),
        }

        if wb.properties:
            if wb.properties.creator:
                metadata["author"] = wb.properties.creator
            if wb.properties.title:
                title = wb.properties.title
            if wb.properties.created:
                metadata["date"] = wb.properties.created.strftime("%Y-%m-%d")

        date_from_name = extract_date_from_filename(path.stem)
        if date_from_name and "date" not in metadata:
            metadata["date"] = date_from_name

        return ParsedDocument(
            doc_id=path.name,
            title=title,
            content=content,
            tables=tables_md,
            metadata=metadata,
        )
